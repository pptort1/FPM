from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from typing import Optional
import datetime, io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..database import get_db
from ..models import Transaccion
from ..clasificador import calcular_iva, PLAN
from ..parser_santander import _firma

router = APIRouter(prefix="/api/egresos", tags=["egresos"])

COLS = [
    ("Fecha Pago",   "DD/MM/AAAA — ej: 15/05/2025",  14),
    ("Mes Devengo",  "AAAA-MM — ej: 2025-05",         12),
    ("Descripcion",  "Glosa del gasto",                40),
    ("Proveedor",    "Nombre proveedor",                30),
    ("Monto Total",  "Monto en pesos (sin puntos)",    16),
    ("Tipo Doc",     "F=Factura  S=Boleta Honorarios B=Boleta", 10),
    ("Forma Pago",   "Credito=TC  Debito=CuentaCorriente  Efectivo=Caja", 14),
    ("Cuenta",       "Código cuenta (ej: 1.1.1) — opcional", 10),
    ("CC",           "Centro costo (ej: CC1) — opcional",      8),
]

EJEMPLO = [
    "15/05/2025", "2025-05", "Arriendo local mayo", "Inversiones XYZ Ltda",
    "500000", "F", "Debito", "1.3.1", "CC3",
]


@router.get("/plantilla")
def descargar_plantilla():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Egresos"

    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT = Font(color="FFFFFF", bold=True)
    GRAY     = PatternFill("solid", fgColor="F2F2F2")
    ITALIC   = Font(color="9E9E9E", italic=True)

    # Fila 1: cabecera
    for j, (col, _, w) in enumerate(COLS, 1):
        c = ws.cell(1, j, col)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width = w

    # Fila 2: ayuda
    for j, (_, hint, _) in enumerate(COLS, 1):
        c = ws.cell(2, j, hint)
        c.fill = GRAY; c.font = ITALIC
        c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Fila 3: ejemplo
    for j, val in enumerate(EJEMPLO, 1):
        c = ws.cell(3, j, val)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_egresos.xlsx"},
    )


# ── Importar Excel llenado ────────────────────────────────────────────────

@router.post("/importar-excel")
def importar_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contenido = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=3, values_only=True))  # salta cabecera + ayuda
    wb.close()

    firmas_db: set[str] = set(
        r[0] for r in db.execute(
            __import__("sqlalchemy", fromlist=["select"]).select(Transaccion.firma_dedup)
            .where(Transaccion.firma_dedup.isnot(None))
        ).all()
    )

    agregados = dupes = errores = 0
    nuevos = []
    for r in rows:
        if not any(r) or not r[0]:
            continue
        try:
            fecha = _parse(r[0])
            mes   = str(r[1] or "")[:7] or f"{fecha.year}-{fecha.month:02d}"
            monto = int(str(r[4] or "0").replace(".", "").replace(",", "").strip())
            tipo  = str(r[5] or "S").strip().upper()[:1]
            tipo  = tipo if tipo in ("F","S","B") else "S"
            forma_raw = str(r[6] or "Debito").strip().lower()
            if "cr" in forma_raw:    forma = "Credito"
            elif "ef" in forma_raw:  forma = "Efectivo"
            else:                    forma = "Debito"
            cuenta= str(r[7] or "").strip()
            cc    = str(r[8] or "").strip()
            if cuenta and cuenta in PLAN:
                nombre_cta, cc = PLAN[cuenta]
            else:
                nombre_cta = str(r[7] or "") or "Sin clasificar"

            iva  = calcular_iva(monto, tipo)
            neto = monto - iva
            desc = str(r[2] or "").strip()
            firma = _firma(fecha, monto, desc)

            if firma in firmas_db:
                dupes += 1; continue

            nuevos.append(Transaccion(
                fecha_pago=fecha, mes_devengo=mes, descripcion=desc,
                proveedor=str(r[3] or "").strip() or None,
                monto_total=monto, tipo_doc=tipo, forma_pago=forma,
                nombre_cuenta=nombre_cta, iva=iva, monto_neto=neto,
                cuenta=cuenta, cc=cc,
                estado="validado", confianza=100, firma_dedup=firma,
            ))
            firmas_db.add(firma)
            agregados += 1
        except Exception:
            errores += 1

    if nuevos:
        db.add_all(nuevos); db.commit()

    return {"agregados": agregados, "duplicados": dupes, "errores": errores}


def _parse(v) -> datetime.date:
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.date() if isinstance(v, datetime.datetime) else v
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(str(v).strip(), fmt).date()
        except Exception:
            pass
    raise ValueError(f"Fecha inválida: {v}")


# ── Egreso manual (un registro) ───────────────────────────────────────────

class EgresoManualIn(BaseModel):
    fecha_pago:   str          # "2025-05-15"
    mes_devengo:  str          # "2025-05"
    descripcion:  str
    proveedor:    Optional[str] = None
    monto_total:  int
    tipo_doc:     str = "S"   # F / S / B
    forma_pago:   str = "Debito"
    cuenta:       str = ""
    cc:           str = ""

@router.post("/manual")
def agregar_manual(body: EgresoManualIn, db: Session = Depends(get_db)):
    fecha = datetime.date.fromisoformat(body.fecha_pago)
    nombre_cta = PLAN[body.cuenta][0] if body.cuenta in PLAN else "Sin clasificar"
    cc = PLAN[body.cuenta][1] if body.cuenta in PLAN else body.cc
    iva  = calcular_iva(body.monto_total, body.tipo_doc)
    neto = body.monto_total - iva
    firma = _firma(fecha, body.monto_total, body.descripcion)

    tx = Transaccion(
        fecha_pago=fecha, mes_devengo=body.mes_devengo,
        descripcion=body.descripcion, proveedor=body.proveedor,
        monto_total=body.monto_total, tipo_doc=body.tipo_doc,
        forma_pago=body.forma_pago, nombre_cuenta=nombre_cta,
        iva=iva, monto_neto=neto, cuenta=body.cuenta, cc=cc,
        estado="validado", confianza=100, firma_dedup=firma,
    )
    db.add(tx); db.commit(); db.refresh(tx)
    return {"id": tx.id, "mensaje": "Egreso registrado"}
