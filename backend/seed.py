"""
seed.py — Inicializa la BD.
- Primera vez: crea tablas e importa historial desde Excel.
- Siguientes veces: solo crea tablas nuevas + columnas nuevas; NO borra datos.
"""
import sys
from pathlib import Path
from datetime import date, datetime
import openpyxl
from sqlalchemy import text, select, func

sys.path.insert(0, str(Path(__file__).parent))
from app.database import engine, SessionLocal, Base
from app.models import Transaccion, Ingreso, Usuario, ProveedorMaestro
from app.auth import hash_password
from app.config import settings
from app.parser_santander import _firma

EXCEL_EGRESOS = Path(__file__).parent / "data" / "egresos_2025_FINAL.xlsx"
EXCEL_PLAN    = Path(__file__).parent / "data" / "plan_cuentas.xlsx"


def _to_date(v) -> date | None:
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    try: return datetime.strptime(str(v), "%d/%m/%Y").date()
    except Exception: return None


def _migrate(conn):
    """Agrega columnas nuevas sin borrar datos existentes."""
    nuevas = [
        ("transacciones", "estado",       "VARCHAR(20) DEFAULT 'validado'"),
        ("transacciones", "firma_dedup",  "VARCHAR(32)"),
        ("transacciones", "rut",          "VARCHAR(20)"),
        ("transacciones", "confianza",    "INTEGER DEFAULT 100"),
        ("transacciones", "reconciliado", "BOOLEAN DEFAULT FALSE"),
        ("ingresos",      "bsale_id",     "VARCHAR(20)"),
    ]
    for tabla, col, tipo in nuevas:
        try:
            conn.execute(text(f"ALTER TABLE {tabla} ADD COLUMN IF NOT EXISTS {col} {tipo}"))
        except Exception:
            # SQLite no soporta IF NOT EXISTS — intenta sin él
            try:
                conn.execute(text(f"ALTER TABLE {tabla} ADD COLUMN {col} {tipo}"))
            except Exception:
                pass  # columna ya existe
    conn.commit()


def main():
    # Crea tablas nuevas SIN borrar las existentes
    Base.metadata.create_all(bind=engine)

    with engine.connect() as conn:
        _migrate(conn)

    with SessionLocal() as db:
        n_tx  = db.execute(select(func.count()).select_from(Transaccion)).scalar_one()
        n_ing = db.execute(select(func.count()).select_from(Ingreso)).scalar_one()
        es_primera_vez = (n_tx == 0 and n_ing == 0)

    if es_primera_vez:
        print("Primera ejecución — importando historial desde Excel...")
        _seed_egresos()
        _seed_ingresos()
    else:
        print(f"BD existente ({n_tx} egresos, {n_ing} ingresos) — omitiendo seed.")

    _ensure_admin()
    print("Listo.")


def _seed_egresos():
    wb = openpyxl.load_workbook(EXCEL_EGRESOS, read_only=True, data_only=True)
    rows = list(wb["EGRESOS 2025"].iter_rows(min_row=2, values_only=True))
    wb.close()

    egresos, seen = [], set()
    for r in rows:
        fecha = _to_date(r[0])
        if not fecha: continue
        firma = _firma(fecha, int(r[4] or 0), str(r[2] or ""))
        if firma in seen: continue
        seen.add(firma)
        egresos.append(Transaccion(
            fecha_pago=fecha, mes_devengo=str(r[1] or "")[:7],
            descripcion=str(r[2] or ""), proveedor=str(r[3]) if r[3] else None,
            monto_total=int(r[4] or 0), tipo_doc=str(r[5] or "S"),
            forma_pago=str(r[6] or "Debito"), nombre_cuenta=str(r[7]) if r[7] else None,
            iva=int(r[8] or 0), monto_neto=int(r[9] or 0),
            cuenta=str(r[10] or ""), cc=str(r[11] or ""),
            estado="validado", confianza=100, firma_dedup=firma,
        ))

    with SessionLocal() as db:
        db.bulk_save_objects(egresos); db.commit()
    print(f"Egresos importados: {len(egresos)}")


def _seed_ingresos():
    wb = openpyxl.load_workbook(EXCEL_PLAN, read_only=True, data_only=True)
    rows = list(wb["1. INGRESOS"].iter_rows(min_row=2, values_only=True))
    wb.close()

    ingresos = []
    for r in rows:
        fecha = _to_date(r[0])
        if not fecha: continue
        ingresos.append(Ingreso(
            fecha=fecha, mes_devengo=str(r[1] or "")[:7],
            cliente=str(r[2]) if r[2] else None, descripcion=str(r[3] or ""),
            monto_total=int(r[4] or 0), tipo_doc=str(r[5] or "B"),
            iva=int(r[6] or 0), monto_neto=int(r[7] or 0),
            cuenta=str(r[8] or ""), nombre_cuenta=str(r[9]) if r[9] else None,
            canal=str(r[10]) if r[10] else None,
        ))

    with SessionLocal() as db:
        db.bulk_save_objects(ingresos); db.commit()
    print(f"Ingresos importados: {len(ingresos)}")


def _ensure_admin():
    with SessionLocal() as db:
        u = db.execute(select(Usuario).where(Usuario.username == "admin")).scalar_one_or_none()
        if not u:
            db.add(Usuario(
                username="admin",
                password_hash=hash_password(settings.ADMIN_PASSWORD),
                nombre="Administrador",
                activo=True,
            ))
            db.commit()
            print("Usuario admin creado.")
        else:
            # Actualizar contraseña si cambió
            u.password_hash = hash_password(settings.ADMIN_PASSWORD)
            db.commit()


if __name__ == "__main__":
    main()
